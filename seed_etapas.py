"""Importa o historico real de prazos/realizacoes da aba ETAPAS 8D do Excel
original, casando cada linha pelo ID MRHB com as reclamacoes ja importadas
(seed_reclamacoes.py).

So preenche reclamacoes que AINDA NAO TEM nenhuma etapa registrada no banco
(normalmente as recem-importadas por seed_reclamacoes.py). Reclamacoes que ja
tem progresso lancado direto no sistema web (pela pagina Etapas 8D) sao
puladas de proposito - rodar este script de novo NUNCA sobrescreve prazo ou
realizacao ja preenchidos pelo uso real do sistema.

Uso: python seed_etapas.py "C:\\caminho\\para\\GERENCIADOR_PDCA_2026.xlsm"
"""
import sys
from datetime import datetime
from pathlib import Path

import openpyxl

from app import ETAPAS_8D
from db import get_connection, init_db

COL_ID_MRHB = 1

# Colunas prazo/realizacao de cada etapa, na mesma ordem de ETAPAS_8D
COLUNAS_ETAPAS = [
    (8, 9),    # Contenção
    (10, 11),  # Causa Raiz
    (12, 13),  # Ação Corretiva
    (14, 15),  # Poka Yoke
    (16, 17),  # LPA
    (18, 19),  # Validação
    (20, 21),  # Trab. Padronizado
    (22, 23),  # FMEA
    (24, 25),  # Plano de Controle
    (26, 27),  # Lições Aprendidas
    (28, 29),  # 8D Fechado
]


def texto_data(valor):
    if valor is None or str(valor).strip() == "":
        return None
    if isinstance(valor, datetime):
        return valor.strftime("%Y-%m-%d")
    return str(valor).strip()


def seed(source_path: str):
    wb = openpyxl.load_workbook(source_path, data_only=True, keep_vba=True)
    ws = wb["ETAPAS 8D"]

    init_db()
    conn = get_connection()

    reclamacoes_por_id_mrhb = {
        row["id_mrhb"]: row["id"]
        for row in conn.execute("SELECT id, id_mrhb FROM reclamacoes").fetchall()
    }
    ja_tem_etapas = {
        row["reclamacao_id"] for row in conn.execute("SELECT DISTINCT reclamacao_id FROM etapas_8d").fetchall()
    }

    novas = 0
    puladas_ja_tem_progresso = 0
    sem_correspondencia = 0

    for r in range(3, ws.max_row + 1):
        id_mrhb = ws.cell(row=r, column=COL_ID_MRHB).value
        if not id_mrhb:
            continue
        id_mrhb = str(id_mrhb).strip()

        reclamacao_id = reclamacoes_por_id_mrhb.get(id_mrhb)
        if reclamacao_id is None:
            sem_correspondencia += 1
            continue

        if reclamacao_id in ja_tem_etapas:
            puladas_ja_tem_progresso += 1
            continue

        for ordem, ((nome, _dias), (col_prazo, col_realizacao)) in enumerate(
            zip(ETAPAS_8D, COLUNAS_ETAPAS), start=1
        ):
            prazo = texto_data(ws.cell(row=r, column=col_prazo).value)
            realizacao = texto_data(ws.cell(row=r, column=col_realizacao).value)

            conn.execute(
                "INSERT INTO etapas_8d (reclamacao_id, etapa, ordem, prazo, realizacao) VALUES (?, ?, ?, ?, ?)",
                (reclamacao_id, nome, ordem, prazo, realizacao),
            )
        novas += 1

    conn.commit()
    conn.close()
    print(
        f"Reclamações novas com etapas importadas: {novas}. "
        f"Puladas (já tinham progresso no sistema): {puladas_ja_tem_progresso}. "
        f"Sem correspondência no banco: {sem_correspondencia}."
    )


if __name__ == "__main__":
    if len(sys.argv) < 2:
        print("Uso: python seed_etapas.py \"C:\\caminho\\para\\GERENCIADOR_PDCA_2026.xlsm\"")
        sys.exit(1)
    source = sys.argv[1]
    if not Path(source).exists():
        print(f"Arquivo nao encontrado: {source}")
        sys.exit(1)
    seed(source)
