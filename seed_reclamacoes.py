"""Importa o historico real de reclamacoes da aba GERENCIADOR do arquivo
Excel original para a tabela `reclamacoes`, preservando o ID MRHB de cada
uma. Isso garante que o historico mostre os dados reais ja existentes e que
a numeracao de novos IDs MRHB continue de onde o Excel parou (a contagem por
mes/ano usada em app.py conta os registros ja existentes na tabela).

Uso: python seed_reclamacoes.py "C:\\caminho\\para\\GERENCIADOR_PDCA_2026.xlsm"
"""
import sys
from datetime import datetime
from pathlib import Path

import openpyxl

from db import get_connection, init_db

COL = {
    "id_mrhb": 6,       # F
    "numero_notificacao": 7,   # G
    "data_abertura": 8,        # H
    "pn_mrhb": 9,               # I
    "descricao_defeito": 10,    # J
    "tipo_pdca": 11,             # K
    "pn_cliente": 12,            # L
    "cliente": 13,                # M
    "area_responsavel_8d": 14,    # N
    "responsavel_cliente": 15,    # O
    "tipo_problema": 16,          # P
    "estratificacao_defeito": 17, # Q
    "tipo_reclamacao": 18,        # R
    "mq_in": 19,                   # S
    "mq_us": 20,                   # T
    "qtd_reclamada": 21,           # U
}


def texto(valor):
    if valor is None:
        return ""
    return str(valor).strip()


def seed(source_path: str):
    wb = openpyxl.load_workbook(source_path, data_only=True, keep_vba=True)
    ws = wb["GERENCIADOR"]

    linhas = []
    for r in range(10, ws.max_row + 1):
        id_mrhb = ws.cell(row=r, column=COL["id_mrhb"]).value
        data_abertura = ws.cell(row=r, column=COL["data_abertura"]).value

        if not id_mrhb or not data_abertura:
            continue

        if isinstance(data_abertura, datetime):
            data_abertura = data_abertura.strftime("%Y-%m-%d")
        else:
            data_abertura = texto(data_abertura)

        linhas.append(
            (
                texto(id_mrhb),
                texto(ws.cell(row=r, column=COL["numero_notificacao"]).value),
                data_abertura,
                texto(ws.cell(row=r, column=COL["pn_mrhb"]).value),
                texto(ws.cell(row=r, column=COL["pn_cliente"]).value),
                texto(ws.cell(row=r, column=COL["descricao_defeito"]).value),
                texto(ws.cell(row=r, column=COL["tipo_pdca"]).value),
                texto(ws.cell(row=r, column=COL["cliente"]).value),
                texto(ws.cell(row=r, column=COL["area_responsavel_8d"]).value),
                texto(ws.cell(row=r, column=COL["responsavel_cliente"]).value),
                texto(ws.cell(row=r, column=COL["tipo_problema"]).value),
                texto(ws.cell(row=r, column=COL["estratificacao_defeito"]).value),
                texto(ws.cell(row=r, column=COL["tipo_reclamacao"]).value),
                texto(ws.cell(row=r, column=COL["mq_in"]).value),
                texto(ws.cell(row=r, column=COL["mq_us"]).value),
                texto(ws.cell(row=r, column=COL["qtd_reclamada"]).value),
            )
        )

    init_db()
    conn = get_connection()
    conn.executemany(
        """
        INSERT INTO reclamacoes (
            id_mrhb, numero_notificacao, data_abertura, pn_mrhb, pn_cliente,
            descricao_defeito, tipo_pdca, cliente, area_responsavel_8d,
            responsavel_cliente, tipo_problema, estratificacao_defeito,
            tipo_reclamacao, mq_in, mq_us, qtd_reclamada
        ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
        ON CONFLICT(id_mrhb) DO NOTHING
        """,
        linhas,
    )
    conn.commit()
    total = conn.execute("SELECT COUNT(*) FROM reclamacoes").fetchone()[0]
    conn.close()
    print(f"Linhas lidas do Excel: {len(linhas)}. Total na tabela reclamacoes: {total}")


if __name__ == "__main__":
    if len(sys.argv) < 2:
        print("Uso: python seed_reclamacoes.py \"C:\\caminho\\para\\GERENCIADOR_PDCA_2026.xlsm\"")
        sys.exit(1)
    source = sys.argv[1]
    if not Path(source).exists():
        print(f"Arquivo nao encontrado: {source}")
        sys.exit(1)
    seed(source)
