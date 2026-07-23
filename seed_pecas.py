"""Importa a lista mestre de pecas (aba LISTA PN) do arquivo Excel original
para a tabela `pecas`, usada para o autopreenchimento de cliente/responsavel
no formulario de nova reclamacao.

Uso: python seed_pecas.py "C:\\caminho\\para\\GERENCIADOR_PDCA_2026.xlsm"
"""
import sys
from pathlib import Path

import openpyxl

from db import get_connection, init_db


def seed(source_path: str):
    wb = openpyxl.load_workbook(source_path, data_only=True, keep_vba=True)
    ws = wb["LISTA PN"]

    rows = []
    for r in range(2, ws.max_row + 1):
        pn_mrhb = ws.cell(row=r, column=1).value
        cliente = ws.cell(row=r, column=2).value
        pn_cliente = ws.cell(row=r, column=3).value
        responsavel = ws.cell(row=r, column=4).value

        if pn_mrhb is None or str(pn_mrhb).strip() == "":
            continue

        pn_mrhb = str(pn_mrhb).strip().replace("\xa0", "")
        cliente = (str(cliente).strip() if cliente is not None else "")
        pn_cliente = (str(pn_cliente).strip() if pn_cliente is not None else "")
        responsavel = (str(responsavel).strip() if responsavel is not None else "")

        rows.append((pn_mrhb, cliente, pn_cliente, responsavel))

    init_db()
    conn = get_connection()
    conn.executemany(
        """
        INSERT INTO pecas (pn_mrhb, cliente, pn_cliente, responsavel)
        VALUES (?, ?, ?, ?)
        ON CONFLICT(pn_mrhb) DO UPDATE SET
            cliente = excluded.cliente,
            pn_cliente = excluded.pn_cliente,
            responsavel = excluded.responsavel
        """,
        rows,
    )
    conn.commit()
    count = conn.execute("SELECT COUNT(*) FROM pecas").fetchone()[0]
    conn.close()
    print(f"Importadas/atualizadas {len(rows)} pecas. Total na tabela: {count}")


if __name__ == "__main__":
    if len(sys.argv) < 2:
        print("Uso: python seed_pecas.py \"C:\\caminho\\para\\GERENCIADOR_PDCA_2026.xlsm\"")
        sys.exit(1)
    source = sys.argv[1]
    if not Path(source).exists():
        print(f"Arquivo nao encontrado: {source}")
        sys.exit(1)
    seed(source)
